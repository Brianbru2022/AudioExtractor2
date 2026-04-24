import json
import shutil
import tempfile
import unittest
import zipfile
from pathlib import Path

from openpyxl import load_workbook

from app.core.config import load_config
from app.services.exports.service import _docx_action_column_widths
from app.utils.files import sha256_for_file
from tests._support import IsolatedWorkerApp


class ExportTests(unittest.TestCase):
    def setUp(self) -> None:
        self.worker = IsolatedWorkerApp(prefix="audio-extractor-export-tests-").start()
        self.client = self.worker.client
        self.services = self.worker.services
        self.config = load_config()
        self.temp_dir = Path(tempfile.mkdtemp(prefix="audio-extractor-export-tests-"))

    def tearDown(self) -> None:
        self.worker.stop()
        shutil.rmtree(self.temp_dir, ignore_errors=True)

    def test_docx_export_uses_reviewed_only_items(self) -> None:
        meeting_id = self._seed_reviewable_meeting()
        response = self.client.post(
            f"/api/v1/meetings/{meeting_id}/exports",
            json={
                "export_profile": "formal_minutes_pack",
                "format": "docx",
                "reviewed_only": True,
                "include_evidence_appendix": True,
                "include_transcript_appendix": False,
                "include_confidence_flags": True,
            },
        )
        self.assertEqual(response.status_code, 200)
        export_run = response.json()
        self.assertEqual(export_run["status"], "completed")
        file_path = Path(export_run["file_path"])
        self.assertTrue(file_path.exists())
        with zipfile.ZipFile(file_path) as archive:
            xml = archive.read("word/document.xml").decode("utf-8")
        self.assertIn("Weekly Operating Review", xml)
        self.assertIn("Finalize the rollout plan", xml)
        self.assertNotIn("Rejected follow-up placeholder", xml)

    def test_docx_minutes_split_inline_section_headings(self) -> None:
        meeting_id = self._seed_reviewable_meeting()
        extraction_run = self.services["extraction_run_repository"].get_latest_for_meeting(meeting_id)
        assert extraction_run is not None
        self.services["extraction_summary_repository"].replace_for_run(
            int(extraction_run["id"]),
            meeting_id,
            "Concise summary",
            (
                "KEY DISCUSSION TOPICS: Kennedys' slow response. "
                "DECISIONS MADE: Push back on the meeting request. "
                "ACTION ITEMS: - Go back to Kennedys. - Request the document. "
                "OPEN QUESTIONS: Will Kennedys reply?"
            ),
        )
        response = self.client.post(
            f"/api/v1/meetings/{meeting_id}/exports",
            json={
                "export_profile": "formal_minutes_pack",
                "format": "docx",
                "reviewed_only": True,
                "include_evidence_appendix": False,
                "include_transcript_appendix": False,
                "include_confidence_flags": False,
            },
        )
        self.assertEqual(response.status_code, 200)
        with zipfile.ZipFile(Path(response.json()["file_path"])) as archive:
            xml = archive.read("word/document.xml").decode("utf-8")
        self.assertIn("KEY DISCUSSION TOPICS", xml)
        self.assertIn("DECISIONS MADE", xml)
        self.assertIn("ACTION ITEMS", xml)
        self.assertIn("OPEN QUESTIONS", xml)
        self.assertIn("Go back to Kennedys.", xml)
        self.assertIn("Request the document.", xml)

    def test_docx_export_includes_attendees_and_circulation(self) -> None:
        meeting_id = self._seed_reviewable_meeting()
        response = self.client.post(
            f"/api/v1/meetings/{meeting_id}/exports",
            json={
                "export_profile": "formal_minutes_pack",
                "format": "docx",
                "reviewed_only": True,
                "include_evidence_appendix": False,
                "include_transcript_appendix": False,
                "include_confidence_flags": False,
            },
        )
        self.assertEqual(response.status_code, 200)
        with zipfile.ZipFile(Path(response.json()["file_path"])) as archive:
            xml = archive.read("word/document.xml").decode("utf-8")
        self.assertIn("Attendees", xml)
        self.assertIn("Circulation", xml)
        self.assertIn("Alex Morgan", xml)
        self.assertIn("Brian Review", xml)
        self.assertIn("10-April-26", xml)
        self.assertNotIn("Prepared from reviewed meeting transcript and evidence-backed outputs.", xml)
        self.assertNotIn("FORMAL MINUTES PACK", xml)
        self.assertNotIn("STATUS", xml)
        self.assertNotIn("GENERATED", xml)

    def test_docx_action_table_uses_priority_without_evidence_or_confidence(self) -> None:
        meeting_id = self._seed_reviewable_meeting()
        response = self.client.post(
            f"/api/v1/meetings/{meeting_id}/exports",
            json={
                "export_profile": "formal_minutes_pack",
                "format": "docx",
                "reviewed_only": True,
                "include_evidence_appendix": False,
                "include_transcript_appendix": False,
                "include_confidence_flags": True,
            },
        )
        self.assertEqual(response.status_code, 200)
        with zipfile.ZipFile(Path(response.json()["file_path"])) as archive:
            xml = archive.read("word/document.xml").decode("utf-8")
        self.assertIn("Priority", xml)
        self.assertNotIn(">Review<", xml)
        self.assertNotIn(">Evidence<", xml)
        self.assertNotIn(">Confidence<", xml)
        self.assertIn("18-April-26", xml)

    def test_docx_action_table_width_fits_printable_page(self) -> None:
        total_width = sum(_docx_action_column_widths())
        printable_width = 8.5 - 0.75 - 0.75
        self.assertLessEqual(total_width, printable_width)

    def test_csv_and_xlsx_action_exports_have_expected_columns(self) -> None:
        meeting_id = self._seed_reviewable_meeting()
        csv_response = self.client.post(
            f"/api/v1/meetings/{meeting_id}/exports",
            json={
                "export_profile": "action_register",
                "format": "csv",
                "reviewed_only": False,
                "include_evidence_appendix": False,
                "include_transcript_appendix": False,
                "include_confidence_flags": True,
            },
        )
        self.assertEqual(csv_response.status_code, 200)
        csv_path = Path(csv_response.json()["file_path"])
        csv_text = csv_path.read_text(encoding="utf-8")
        self.assertIn("action_text,owner,due_date,review_status,explicit_or_inferred,evidence_timestamps,confidence", csv_text)
        self.assertIn("Finalize the rollout plan", csv_text)
        self.assertIn("Rejected follow-up placeholder", csv_text)

        xlsx_response = self.client.post(
            f"/api/v1/meetings/{meeting_id}/exports",
            json={
                "export_profile": "action_register",
                "format": "xlsx",
                "reviewed_only": True,
                "include_evidence_appendix": False,
                "include_transcript_appendix": False,
                "include_confidence_flags": True,
            },
        )
        self.assertEqual(xlsx_response.status_code, 200)
        workbook = load_workbook(Path(xlsx_response.json()["file_path"]))
        sheet = workbook.active
        headers = [cell.value for cell in sheet[1]]
        self.assertEqual(
            headers,
            ["Action Text", "Owner", "Due Date", "Review Status", "Explicit Or Inferred", "Evidence Timestamps", "Confidence"],
        )
        values = [sheet.cell(row=2, column=1).value]
        self.assertIn("Finalize the rollout plan", values)

    def test_json_archive_contains_expected_structure(self) -> None:
        meeting_id = self._seed_reviewable_meeting()
        response = self.client.post(
            f"/api/v1/meetings/{meeting_id}/exports",
            json={
                "export_profile": "full_archive",
                "format": "json",
                "reviewed_only": True,
                "include_evidence_appendix": True,
                "include_transcript_appendix": True,
                "include_confidence_flags": True,
            },
        )
        self.assertEqual(response.status_code, 200)
        payload = json.loads(Path(response.json()["file_path"]).read_text(encoding="utf-8"))
        self.assertEqual(payload["meeting"]["id"], meeting_id)
        self.assertIn("merged_transcript", payload)
        self.assertIn("extraction", payload)
        self.assertIn("summary", payload["extraction"])
        self.assertEqual(len(payload["extraction"]["actions"]), 1)

    def test_export_can_write_to_custom_output_directory(self) -> None:
        meeting_id = self._seed_reviewable_meeting()
        custom_output_dir = self.temp_dir / "chosen-exports"
        response = self.client.post(
            f"/api/v1/meetings/{meeting_id}/exports",
            json={
                "export_profile": "transcript_export",
                "format": "txt",
                "reviewed_only": True,
                "include_evidence_appendix": False,
                "include_transcript_appendix": False,
                "include_confidence_flags": False,
                "output_directory": str(custom_output_dir),
            },
        )
        self.assertEqual(response.status_code, 200)
        export_run = response.json()
        file_path = Path(export_run["file_path"])
        self.assertTrue(file_path.exists())
        self.assertEqual(file_path.parent, custom_output_dir)
        self.assertEqual(file_path.name, "Operations - 2026-04-10 - 01.txt")

    def test_export_file_names_use_project_date_and_sequence(self) -> None:
        meeting_id = self._seed_reviewable_meeting()
        first = self.client.post(
            f"/api/v1/meetings/{meeting_id}/exports",
            json={
                "export_profile": "transcript_export",
                "format": "txt",
                "reviewed_only": True,
                "include_evidence_appendix": False,
                "include_transcript_appendix": False,
                "include_confidence_flags": False,
            },
        )
        self.assertEqual(first.status_code, 200)
        second = self.client.post(
            f"/api/v1/meetings/{meeting_id}/exports",
            json={
                "export_profile": "formal_minutes_pack",
                "format": "docx",
                "reviewed_only": True,
                "include_evidence_appendix": False,
                "include_transcript_appendix": False,
                "include_confidence_flags": False,
            },
        )
        self.assertEqual(second.status_code, 200)

        first_name = Path(first.json()["file_path"]).name
        second_name = Path(second.json()["file_path"]).name
        self.assertEqual(first_name, "Operations - 2026-04-10 - 01.txt")
        self.assertEqual(second_name, "Operations - 2026-04-10 - 02.docx")

    def test_export_history_persists_completed_and_failed_runs(self) -> None:
        good_meeting = self._seed_reviewable_meeting()
        good = self.client.post(
            f"/api/v1/meetings/{good_meeting}/exports",
            json={
                "export_profile": "transcript_export",
                "format": "txt",
                "reviewed_only": True,
                "include_evidence_appendix": False,
                "include_transcript_appendix": False,
                "include_confidence_flags": False,
            },
        )
        self.assertEqual(good.status_code, 200)

        bad_meeting, _ = self._seed_transcript_only_meeting()
        bad = self.client.post(
            f"/api/v1/meetings/{bad_meeting}/exports",
            json={
                "export_profile": "formal_minutes_pack",
                "format": "pdf",
                "reviewed_only": True,
                "include_evidence_appendix": True,
                "include_transcript_appendix": False,
                "include_confidence_flags": False,
            },
        )
        self.assertEqual(bad.status_code, 400)

        good_history = self.client.get(f"/api/v1/meetings/{good_meeting}/exports")
        self.assertEqual(good_history.status_code, 200)
        self.assertEqual(good_history.json()[0]["status"], "completed")

        bad_history = self.client.get(f"/api/v1/meetings/{bad_meeting}/exports")
        self.assertEqual(bad_history.status_code, 200)
        self.assertEqual(bad_history.json()[0]["status"], "failed")
        self.assertIn("completed extraction summary", bad_history.json()[0]["error_message"])

    def test_txt_transcript_export_is_readable(self) -> None:
        meeting_id = self._seed_reviewable_meeting()
        response = self.client.post(
            f"/api/v1/meetings/{meeting_id}/exports",
            json={
                "export_profile": "transcript_export",
                "format": "txt",
                "reviewed_only": True,
                "include_evidence_appendix": False,
                "include_transcript_appendix": False,
                "include_confidence_flags": False,
            },
        )
        self.assertEqual(response.status_code, 200)
        text = Path(response.json()["file_path"]).read_text(encoding="utf-8")
        self.assertIn("[00:00:00 - 00:00:05] Speaker 1: We agreed to finalize the rollout plan by next Friday.", text)

    def _seed_reviewable_meeting(self) -> int:
        meeting_id, transcription_run_id = self._seed_transcript_only_meeting()

        extraction_run_id = self.services["extraction_run_repository"].create(
            meeting_id=meeting_id,
            transcription_run_id=transcription_run_id,
            job_run_id=self.services["job_run_repository"].create(meeting_id, "extract", "seed extraction"),
            model="gemini-3.1-pro-preview",
            model_version="gemini-3.1-pro-preview",
            config_json={"seeded": True},
        )
        action_ids = self.services["extracted_action_repository"].replace_for_run(
            extraction_run_id,
            meeting_id,
            [
                {
                    "text": "Finalize the rollout plan",
                    "owner": "Alex",
                    "due_date": "2026-04-18",
                    "priority": "high",
                    "confidence": 0.94,
                    "explicit_or_inferred": "explicit",
                    "review_status": "accepted",
                },
                {
                    "text": "Rejected follow-up placeholder",
                    "owner": None,
                    "due_date": None,
                    "priority": None,
                    "confidence": 0.52,
                    "explicit_or_inferred": "inferred",
                    "review_status": "rejected",
                },
            ],
        )
        decision_ids = self.services["extracted_decision_repository"].replace_for_run(
            extraction_run_id,
            meeting_id,
            [
                {
                    "text": "Proceed with the phased rollout.",
                    "confidence": 0.91,
                    "explicit_or_inferred": "explicit",
                    "review_status": "accepted",
                }
            ],
        )
        risk_ids = self.services["extracted_risk_repository"].replace_for_run(
            extraction_run_id,
            meeting_id,
            [
                {
                    "text": "Vendor delivery may slip by one week.",
                    "confidence": 0.85,
                    "explicit_or_inferred": "explicit",
                    "review_status": "accepted",
                }
            ],
        )
        question_ids = self.services["extracted_question_repository"].replace_for_run(
            extraction_run_id,
            meeting_id,
            [
                {
                    "text": "Can support cover the launch weekend?",
                    "confidence": 0.79,
                    "explicit_or_inferred": "explicit",
                    "review_status": "accepted",
                }
            ],
        )
        self.services["extracted_topic_repository"].replace_for_run(
            extraction_run_id,
            meeting_id,
            [
                {
                    "text": "Rollout readiness and owner alignment.",
                    "confidence": 0.86,
                    "explicit_or_inferred": "explicit",
                    "review_status": "accepted",
                }
            ],
        )
        self.services["extraction_evidence_repository"].replace_for_run(
            extraction_run_id,
            [
                {
                    "entity_type": "action",
                    "entity_id": action_ids[0],
                    "transcript_segment_id": 1,
                    "start_ms": 0,
                    "end_ms": 5000,
                    "speaker_label": "Speaker 1",
                    "quote_snippet": "We agreed to finalize the rollout plan by next Friday.",
                    "confidence": 0.95,
                },
                {
                    "entity_type": "decision",
                    "entity_id": decision_ids[0],
                    "transcript_segment_id": 1,
                    "start_ms": 0,
                    "end_ms": 5000,
                    "speaker_label": "Speaker 1",
                    "quote_snippet": "We agreed to finalize the rollout plan by next Friday.",
                    "confidence": 0.9,
                },
                {
                    "entity_type": "risk",
                    "entity_id": risk_ids[0],
                    "transcript_segment_id": 2,
                    "start_ms": 6000,
                    "end_ms": 10000,
                    "speaker_label": "Speaker 2",
                    "quote_snippet": "Vendor delivery may slip by one week.",
                    "confidence": 0.85,
                },
                {
                    "entity_type": "question",
                    "entity_id": question_ids[0],
                    "transcript_segment_id": 3,
                    "start_ms": 11000,
                    "end_ms": 14000,
                    "speaker_label": "Speaker 3",
                    "quote_snippet": "Can support cover the launch weekend?",
                    "confidence": 0.78,
                },
            ],
        )
        self.services["extraction_summary_repository"].replace_for_run(
            extraction_run_id,
            meeting_id,
            "The team aligned on a phased rollout and captured one high-priority action.",
            "The meeting confirmed a phased rollout approach. Alex will finalize the rollout plan by next Friday. The team noted a vendor delivery risk and raised a support coverage question.",
        )
        self.services["extraction_run_repository"].finalize_success(extraction_run_id)
        return meeting_id

    def _seed_transcript_only_meeting(self) -> tuple[int, int]:
        meeting_id = self.services["meeting_repository"].create(
            "Weekly Operating Review",
            "2026-04-10",
            "Operations",
            "Seeded export test meeting",
            ["Alex Morgan", "Casey Smith"],
            ["Brian Review", "Operations Board"],
        )
        source_path = self.temp_dir / f"meeting_{meeting_id}.wav"
        source_path.write_bytes(b"source")
        self.services["source_file_repository"].create(
            {
                "meeting_id": meeting_id,
                "import_mode": "reference",
                "original_path": str(source_path),
                "managed_copy_path": None,
                "normalized_audio_path": str(source_path),
                "file_name": source_path.name,
                "extension": ".wav",
                "mime_type": "audio/wav",
                "media_type": "audio",
                "size_bytes": source_path.stat().st_size,
                "sha256": sha256_for_file(source_path),
                "duration_ms": 14_000,
                "sample_rate": 16_000,
                "channels": 1,
            }
        )
        self.services["meeting_repository"].update_status(meeting_id, "transcribed")
        preprocessing_job_run_id = self.services["job_run_repository"].create(meeting_id, "preprocess", "seed preprocess")
        preprocessing_run_id = self.services["run_repository"].create(
            meeting_id, preprocessing_job_run_id, self.config.worker_version, {"seeded": True}
        )
        self.services["run_repository"].finalize_success(
            preprocessing_run_id,
            normalized_format="flac",
            normalized_sample_rate=16_000,
            normalized_channels=1,
            silence_map={"candidate_count": 0},
            chunk_strategy={"coverage_validation": {"covers_full_duration": True}},
            waveform_summary={"buckets": [0.1, 0.2, 0.15]},
        )
        transcription_job_run_id = self.services["job_run_repository"].create(meeting_id, "transcribe", "seed transcribe")
        transcription_run_id = self.services["transcription_run_repository"].create(
            meeting_id=meeting_id,
            preprocessing_run_id=preprocessing_run_id,
            job_run_id=transcription_job_run_id,
            engine="google_speech_to_text_v2",
            engine_model="chirp_3",
            language_code="en-US",
            diarization_enabled=True,
            automatic_punctuation_enabled=True,
            chunk_count=1,
            config_json={"seeded": True},
        )
        self.services["transcription_run_repository"].update_progress(
            transcription_run_id,
            completed_chunk_count=1,
            failed_chunk_count=0,
        )
        self.services["transcript_segment_repository"].replace_for_run(
            transcription_run_id,
            "merged",
            [
                {
                    "meeting_id": meeting_id,
                    "chunk_id": None,
                    "segment_index": 0,
                    "speaker_label": "Speaker 1",
                    "speaker_name": None,
                    "text": "We agreed to finalize the rollout plan by next Friday.",
                    "start_ms_in_meeting": 0,
                    "end_ms_in_meeting": 5000,
                    "start_ms_in_chunk": None,
                    "end_ms_in_chunk": None,
                    "confidence": 0.95,
                    "source_type": "merged",
                },
                {
                    "meeting_id": meeting_id,
                    "chunk_id": None,
                    "segment_index": 1,
                    "speaker_label": "Speaker 2",
                    "speaker_name": None,
                    "text": "Vendor delivery may slip by one week.",
                    "start_ms_in_meeting": 6000,
                    "end_ms_in_meeting": 10000,
                    "start_ms_in_chunk": None,
                    "end_ms_in_chunk": None,
                    "confidence": 0.84,
                    "source_type": "merged",
                },
                {
                    "meeting_id": meeting_id,
                    "chunk_id": None,
                    "segment_index": 2,
                    "speaker_label": "Speaker 3",
                    "speaker_name": None,
                    "text": "Can support cover the launch weekend?",
                    "start_ms_in_meeting": 11000,
                    "end_ms_in_meeting": 14000,
                    "start_ms_in_chunk": None,
                    "end_ms_in_chunk": None,
                    "confidence": 0.79,
                    "source_type": "merged",
                },
            ],
        )
        self.services["transcription_run_repository"].finalize_success(transcription_run_id, average_confidence=0.86)
        return meeting_id, transcription_run_id


if __name__ == "__main__":
    unittest.main()
